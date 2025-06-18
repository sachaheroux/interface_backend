import pandas as pd
import io
from typing import Dict, List, Tuple, Optional
from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def parse_flowshop_excel(file_content: bytes) -> Dict:
    """
    Parse un fichier Excel pour les algorithmes flowshop (SPT, EDD, etc.)
    Supporte deux formats:
    1. Format avec onglets (ancien)
    2. Format matrice unique (nouveau - 12 colonnes x 11 lignes)
    
    Args:
        file_content: Contenu du fichier Excel en bytes
        
    Returns:
        Dict contenant les données formatées pour l'API
    """
    try:
        # Essayer d'abord le nouveau format (matrice unique)
        try:
            return parse_matrix_format(file_content)
        except:
            pass
        
        # Fallback vers l'ancien format avec onglets
        excel_file = pd.ExcelFile(io.BytesIO(file_content))
        
        # Vérifier que les onglets requis existent
        required_sheets = ['Machines', 'Jobs']
        missing_sheets = [sheet for sheet in required_sheets if sheet not in excel_file.sheet_names]
        
        if missing_sheets:
            raise HTTPException(
                status_code=400, 
                detail=f"Format Excel non reconnu. Utilisez le template fourni."
            )
        
        # Lire les onglets
        machines_df = pd.read_excel(excel_file, sheet_name='Machines')
        jobs_df = pd.read_excel(excel_file, sheet_name='Jobs')
        
        # Parser les machines
        machine_names, machine_mapping = parse_machines(machines_df)
        
        # Parser les jobs
        jobs_data, due_dates, job_names = parse_jobs(jobs_df, machine_names)
        
        return {
            "jobs_data": jobs_data,
            "due_dates": due_dates,
            "job_names": job_names,
            "machine_names": machine_names,
            "unite": "heures"  # Par défaut
        }
        
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=400, detail=f"Erreur lors de la lecture du fichier Excel: {str(e)}")

def parse_matrix_format(file_content: bytes) -> Dict:
    """
    Parse le nouveau format matrice (12 colonnes x 11 lignes)
    Structure:
    - Cellule C5: "Job" (coin haut gauche du tableau)
    - Colonne C (lignes 6-15): noms des jobs
    - Colonnes D-M: durées sur les machines (10 machines max)
    - Colonne N: dates dues
    - Cellule C20: unité de temps (j/h/m)
    
    Args:
        file_content: Contenu du fichier Excel en bytes
        
    Returns:
        Dict contenant les données formatées pour l'API
    """
    try:
        # Lire le fichier Excel sans en-tête automatique
        excel_file = io.BytesIO(file_content)
        df = pd.read_excel(excel_file, header=None)
        
        # Vérifier la structure minimale
        if df.shape[0] < 20 or df.shape[1] < 14:
            raise ValueError("Structure de fichier incorrecte")
        
        # Vérifier que c'est bien le bon format (cellule C5 doit contenir "Job")
        job_header = df.iloc[4, 2]  # C5 (ligne 5, colonne C)
        if pd.isna(job_header) or str(job_header).strip().lower() != "job":
            raise ValueError("Format non reconnu - cellule C5 doit contenir 'Job'")
        
        # Extraire l'unité de temps (cellule C20)
        unite = "heures"  # valeur par défaut
        try:
            unite_cell = df.iloc[19, 2]  # C20
            if pd.notna(unite_cell):
                unite_str = str(unite_cell).lower().strip()
                if unite_str == 'j':
                    unite = "jours"
                elif unite_str == 'h':
                    unite = "heures"
                elif unite_str == 'm':
                    unite = "minutes"
        except:
            pass
        
        # Extraire les données des jobs
        job_names = []
        jobs_data = []
        due_dates = []
        errors = []
        
        # Parcourir les lignes 6-15 (index 5-14)
        for i in range(5, 15):
            row_num = i + 1  # Numéro de ligne Excel (1-indexé)
            
            # Nom du job (colonne C)
            job_name = df.iloc[i, 2]
            if pd.isna(job_name) or not str(job_name).strip():
                continue  # Ligne vide, on passe
            
            job_name = str(job_name).strip()
            
            # Vérifier les durées sur les machines (colonnes D-M, index 3-12)
            job_durations = []
            duration_errors = []
            
            for j in range(3, 13):  # colonnes D à M
                col_letter = chr(68 + j - 3)  # D, E, F, G, H, I, J, K, L, M
                duration = df.iloc[i, j]
                
                if pd.notna(duration) and str(duration).strip():
                    try:
                        duration_val = float(duration)
                        if duration_val < 0:
                            duration_errors.append(f"Durée négative en {col_letter}{row_num}: {duration}")
                        elif duration_val == 0:
                            # Durée nulle acceptée mais pas ajoutée
                            pass
                        else:
                            job_durations.append([j-3, duration_val])  # [machine_id, duration]
                    except (ValueError, TypeError):
                        duration_errors.append(f"Valeur invalide en {col_letter}{row_num}: '{duration}' (doit être un nombre)")
            
            # Vérifier la date due (colonne N, index 13)
            due_date = df.iloc[i, 13]
            due_date_val = 10.0  # valeur par défaut
            
            if pd.notna(due_date) and str(due_date).strip():
                try:
                    due_date_val = float(due_date)
                    if due_date_val < 0:
                        errors.append(f"Date due négative pour '{job_name}' en N{row_num}: {due_date}")
                        due_date_val = 10.0
                except (ValueError, TypeError):
                    errors.append(f"Date due invalide pour '{job_name}' en N{row_num}: '{due_date}' (doit être un nombre)")
                    due_date_val = 10.0
            
            # Ajouter les erreurs de durée si il y en a
            if duration_errors:
                errors.extend([f"Job '{job_name}': {err}" for err in duration_errors])
            
            # Ajouter le job seulement s'il a au moins une durée valide
            if job_durations:
                job_names.append(job_name)
                jobs_data.append(job_durations)
                due_dates.append(due_date_val)
            elif not duration_errors:  # Pas d'erreurs mais pas de durées non plus
                errors.append(f"Job '{job_name}' (ligne {row_num}): aucune durée valide trouvée")
        
        # Si il y a des erreurs, les signaler
        if errors:
            error_msg = "Erreurs dans le fichier Excel:\n" + "\n".join(f"• {err}" for err in errors[:10])
            if len(errors) > 10:
                error_msg += f"\n... et {len(errors) - 10} autres erreurs"
            raise HTTPException(status_code=400, detail=error_msg)
        
        if not job_names:
            raise HTTPException(
                status_code=400,
                detail="Aucun job valide trouvé. Vérifiez que vous avez rempli les noms de jobs et au moins une durée par job."
            )
        
        # Lire les noms des machines depuis les en-têtes (ligne 5, colonnes D-M)
        machine_names = []
        for j in range(3, 13):  # colonnes D à M (index 3-12)
            try:
                header = df.iloc[4, j]  # ligne 5 (index 4)
                if pd.notna(header) and str(header).strip():
                    machine_names.append(str(header).strip())
                else:
                    machine_names.append(f"Machine_{j-3}")
            except:
                machine_names.append(f"Machine_{j-3}")
        
        # Déterminer le nombre de machines réellement utilisées
        max_machine_id = 0
        for job in jobs_data:
            for machine_id, _ in job:
                max_machine_id = max(max_machine_id, machine_id)
        
        # Garder seulement les noms des machines utilisées
        machine_names = machine_names[:max_machine_id + 1]
        
        return {
            "jobs_data": jobs_data,
            "due_dates": due_dates,
            "job_names": job_names,
            "machine_names": machine_names,
            "unite": unite
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise ValueError(f"Erreur parsing format matrice: {str(e)}")

def parse_machines(machines_df: pd.DataFrame) -> Tuple[List[str], Dict[int, str]]:
    """
    Parse l'onglet Machines
    
    Returns:
        Tuple[List[str], Dict[int, str]]: (noms des machines, mapping ID->nom)
    """
    # Vérifier les colonnes requises
    required_columns = ['ID_Machine', 'Nom_Machine']
    missing_columns = [col for col in required_columns if col not in machines_df.columns]
    
    if missing_columns:
        raise HTTPException(
            status_code=400,
            detail=f"Colonnes manquantes dans l'onglet Machines: {', '.join(missing_columns)}"
        )
    
    # Nettoyer les données
    machines_df = machines_df.dropna(subset=['ID_Machine', 'Nom_Machine'])
    
    if machines_df.empty:
        raise HTTPException(status_code=400, detail="Aucune machine valide trouvée dans l'onglet Machines")
    
    # Vérifier que les ID sont des entiers consécutifs commençant à 0
    machine_ids = sorted(machines_df['ID_Machine'].astype(int).tolist())
    expected_ids = list(range(len(machine_ids)))
    
    if machine_ids != expected_ids:
        raise HTTPException(
            status_code=400,
            detail=f"Les ID des machines doivent être des entiers consécutifs commençant à 0. Trouvé: {machine_ids}, Attendu: {expected_ids}"
        )
    
    # Créer le mapping et la liste des noms
    machine_mapping = {}
    machine_names = [''] * len(machine_ids)
    
    for _, row in machines_df.iterrows():
        machine_id = int(row['ID_Machine'])
        machine_name = str(row['Nom_Machine']).strip()
        
        if not machine_name or machine_name.lower() in ['nan', 'none', '[à remplir]']:
            raise HTTPException(
                status_code=400,
                detail=f"Nom de machine manquant pour l'ID {machine_id}"
            )
        
        machine_mapping[machine_id] = machine_name
        machine_names[machine_id] = machine_name
    
    return machine_names, machine_mapping

def parse_jobs(jobs_df: pd.DataFrame, machine_names: List[str]) -> Tuple[List[List[List[float]]], List[float], List[str]]:
    """
    Parse l'onglet Jobs
    
    Returns:
        Tuple: (jobs_data, due_dates, job_names)
    """
    # Vérifier les colonnes de base
    base_columns = ['Nom_Job', 'Date_Echeance']
    missing_base = [col for col in base_columns if col not in jobs_df.columns]
    
    if missing_base:
        raise HTTPException(
            status_code=400,
            detail=f"Colonnes manquantes dans l'onglet Jobs: {', '.join(missing_base)}"
        )
    
    # Vérifier les colonnes des machines (Machine_0, Machine_1, etc.)
    machine_columns = [f'Machine_{i}' for i in range(len(machine_names))]
    missing_machines = [col for col in machine_columns if col not in jobs_df.columns]
    
    if missing_machines:
        raise HTTPException(
            status_code=400,
            detail=f"Colonnes de machines manquantes dans l'onglet Jobs: {', '.join(missing_machines)}"
        )
    
    # Nettoyer les données
    jobs_df = jobs_df.dropna(subset=['Nom_Job', 'Date_Echeance'])
    
    if jobs_df.empty:
        raise HTTPException(status_code=400, detail="Aucun job valide trouvé dans l'onglet Jobs")
    
    jobs_data = []
    due_dates = []
    job_names = []
    
    for index, row in jobs_df.iterrows():
        # Nom du job
        job_name = str(row['Nom_Job']).strip()
        if not job_name or job_name.lower() in ['nan', 'none', '[à remplir]']:
            raise HTTPException(
                status_code=400,
                detail=f"Nom de job manquant à la ligne {index + 2}"
            )
        
        # Date d'échéance
        try:
            due_date = float(row['Date_Echeance'])
            if due_date < 0:
                raise ValueError("Date d'échéance négative")
        except (ValueError, TypeError):
            raise HTTPException(
                status_code=400,
                detail=f"Date d'échéance invalide pour le job '{job_name}' à la ligne {index + 2}"
            )
        
        # Durées des tâches
        job_tasks = []
        for machine_id, machine_name in enumerate(machine_names):
            machine_col = f'Machine_{machine_id}'
            
            try:
                duration = float(row[machine_col])
                if duration < 0:
                    raise ValueError("Durée négative")
                
                # Format: [machine_id, duration]
                job_tasks.append([machine_id, duration])
                
            except (ValueError, TypeError):
                raise HTTPException(
                    status_code=400,
                    detail=f"Durée invalide pour le job '{job_name}' sur la machine '{machine_name}' à la ligne {index + 2}"
                )
        
        jobs_data.append(job_tasks)
        due_dates.append(due_date)
        job_names.append(job_name)
    
    return jobs_data, due_dates, job_names

def create_flowshop_template(template_type: str = "exemple") -> bytes:
    """
    Crée un template Excel pour les algorithmes flowshop
    
    Args:
        template_type: "exemple" ou "vide"
        
    Returns:
        bytes: Contenu du fichier Excel
    """
    # Créer un BytesIO pour le fichier Excel
    output = io.BytesIO()
    
    # Préparer les données
    if template_type == "exemple":
        machines_data = {
            'ID_Machine': [0, 1, 2],
            'Nom_Machine': ['Découpe', 'Assemblage', 'Finition']
        }
        jobs_data = {
            'Nom_Job': ['Job_A', 'Job_B', 'Job_C'],
            'Date_Echeance': [12, 15, 18],
            'Machine_0': [4, 3, 5],
            'Machine_1': [2, 4, 2],
            'Machine_2': [3, 2, 4]
        }
    else:
        machines_data = {
            'ID_Machine': [0, 1, 2],
            'Nom_Machine': ['[À remplir]', '[À remplir]', '[À remplir]']
        }
        jobs_data = {
            'Nom_Job': ['[À remplir]', '[À remplir]', '[À remplir]'],
            'Date_Echeance': ['[À remplir]', '[À remplir]', '[À remplir]'],
            'Machine_0': ['[À remplir]', '[À remplir]', '[À remplir]'],
            'Machine_1': ['[À remplir]', '[À remplir]', '[À remplir]'],
            'Machine_2': ['[À remplir]', '[À remplir]', '[À remplir]']
        }
    
    # Créer les DataFrames
    machines_df = pd.DataFrame(machines_data)
    jobs_df = pd.DataFrame(jobs_data)
    
    # Créer un workbook avec openpyxl pour un contrôle précis de la structure
    wb = Workbook()
    ws = wb.active
    ws.title = "Données"
    
    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    
    # STRUCTURE FIXE 12x12 - TOUJOURS utiliser ces positions exactes
    
    # C5: "JOB" (coin supérieur gauche)
    ws['C5'] = "JOB"
    ws['C5'].font = header_font
    ws['C5'].fill = header_fill
    ws['C5'].alignment = Alignment(horizontal="center")
    
    # C6-C16: Noms des jobs (colonne C = 3, lignes 6 à 16 = max 11 jobs)
    for i in range(11):  # TOUJOURS 11 lignes (C6 à C16)
        if i < len(job_names):
            job_name = job_names[i]
        else:
            job_name = ""  # Cellule vide si pas assez de jobs
        
        cell = ws.cell(row=6+i, column=3, value=job_name)  # Colonne C = 3
    
    # D5-M5: Noms des machines (ligne 5, colonnes D=4 à M=13 = max 10 machines)
    for i in range(10):  # TOUJOURS 10 colonnes (D à M)
        if i < len(machine_names):
            machine_name = machine_names[i]
        else:
            machine_name = ""  # Cellule vide si pas assez de machines
            
        cell = ws.cell(row=5, column=4+i, value=machine_name)  # Colonnes D=4 à M=13
    
    # N5: "Due Date" (TOUJOURS colonne N = 14)
    ws.cell(row=5, column=14, value="Due Date")  # Colonne N = 14
    
    # N6-N16: Dates d'échéance (colonne N=14, lignes 6 à 16)
    for i in range(11):  # TOUJOURS 11 lignes (N6 à N16)
        if i < len(due_dates):
            due_date = due_dates[i]
        else:
            due_date = ""  # Cellule vide si pas assez de dates
            
        cell = ws.cell(row=6+i, column=14, value=due_date)  # Colonne N = 14
    
    # D6-M16: Matrice des temps de traitement (colonnes D=4 à M=13, lignes 6 à 16)
    for job_idx in range(11):  # TOUJOURS 11 lignes de jobs
        for machine_idx in range(10):  # TOUJOURS 10 colonnes de machines
            if job_idx < len(jobs_data) and machine_idx < len(jobs_data[job_idx]):
                duration = jobs_data[job_idx][machine_idx][1]
            else:
                duration = ""  # Cellule vide si pas de données
                
            cell = ws.cell(row=6+job_idx, column=4+machine_idx, value=duration)
    
    # Ajuster la largeur des colonnes (C à N)
    for col in range(3, 15):  # De C=3 à N=14
        column_letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[column_letter].width = 12
    
    # Instructions dans une nouvelle feuille
    instructions_ws = wb.create_sheet("Instructions")
    instructions_ws.append(["Ce template utilise TOUJOURS une structure fixe de 12x12:"])
    instructions_ws.append(["- C5: 'JOB' (coin supérieur gauche)"])
    instructions_ws.append(["- C6-C16: Noms des jobs (11 lignes maximum)"])
    instructions_ws.append(["- D5-M5: Noms des machines (10 colonnes maximum)"])
    instructions_ws.append(["- N5: 'Due Date' (TOUJOURS en colonne N)"])
    instructions_ws.append(["- N6-N16: Dates d'échéance"])
    instructions_ws.append(["- D6-M16: Matrice des temps de traitement"])
    instructions_ws.append(["- C19: 'Unité de temps'"])
    instructions_ws.append(["- C20: Unité (j/h/m selon l'interface)"])
    instructions_ws.append([""])
    instructions_ws.append(["Les cellules vides sont autorisées si moins de 10 machines ou 11 jobs."])
    
    # C19: "Unité de temps"
    ws.cell(row=19, column=3, value="Unité de temps")
    
    # C20: Unité de temps (j/h/m selon l'interface)
    unit_mapping = {
        "jours": "j",
        "heures": "h", 
        "minutes": "m",
        "jour": "j",
        "heure": "h",
        "minute": "m"
    }
    
    # Conversion de l'unité reçue en abréviation
    unit_abbrev = unit_mapping.get(unite.lower(), unite.lower())
    ws.cell(row=20, column=3, value=unit_abbrev)
    
    print(f"DEBUG - Unité de temps ajoutée: '{unite}' -> '{unit_abbrev}' en C20")

    # Sauvegarder le fichier
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

def parse_jobshop_excel(file_content: bytes) -> Dict:
    """
    Parse un fichier Excel pour les algorithmes Jobshop (SPT, EDD, Contraintes)
    Format spécifique avec cellules (séquence, temps)
    
    Args:
        file_content: Contenu du fichier Excel en bytes
        
    Returns:
        Dict contenant les données formatées pour l'API Jobshop
    """
    try:
        # Lire le fichier Excel
        excel_file = io.BytesIO(file_content)
        df = pd.read_excel(excel_file, header=None)
        
        # Vérifier la structure minimale
        if df.shape[0] < 20 or df.shape[1] < 14:
            raise ValueError("Structure de fichier incorrecte")
        
        # Vérifier que c'est bien le format Jobshop (cellule C5 doit contenir "Job")
        job_header = df.iloc[4, 2]  # C5 (ligne 5, colonne C)
        if pd.isna(job_header) or str(job_header).strip().lower() != "job":
            raise ValueError("Format non reconnu - cellule C5 doit contenir 'Job'")
        
        # Extraire l'unité de temps (cellule C20)
        unite = "heures"  # valeur par défaut
        try:
            unite_cell = df.iloc[19, 2]  # C20
            if pd.notna(unite_cell):
                unite_str = str(unite_cell).lower().strip()
                if unite_str == 'j':
                    unite = "jours"
                elif unite_str == 'h':
                    unite = "heures"
                elif unite_str == 'm':
                    unite = "minutes"
        except:
            pass
        
        # Extraire les données des jobs
        job_names = []
        jobs_data = []
        due_dates = []
        errors = []
        
        # Parcourir les lignes 6-15 (index 5-14) pour les jobs
        for i in range(5, 15):
            row_num = i + 1  # Numéro de ligne Excel (1-indexé)
            
            # Nom du job (colonne C)
            job_name = df.iloc[i, 2]
            if pd.isna(job_name) or not str(job_name).strip():
                continue  # Ligne vide, on passe
            
            job_name = str(job_name).strip()
            
            # Extraire les tâches du job (colonnes D-M, index 3-12)
            job_tasks = []
            task_errors = []
            
            for j in range(3, 13):  # colonnes D à M
                col_letter = chr(68 + j - 3)  # D, E, F, G, H, I, J, K, L, M
                cell_value = df.iloc[i, j]
                
                if pd.notna(cell_value) and str(cell_value).strip():
                    try:
                        # Parser le format séquence, temps (avec ou sans parenthèses)
                        cell_str = str(cell_value).strip()
                        
                        # Enlever les parenthèses si présentes
                        if cell_str.startswith('(') and cell_str.endswith(')'):
                            cell_str = cell_str[1:-1]
                        
                        # Séparer par la virgule
                        parts = cell_str.split(',')
                        if len(parts) != 2:
                            task_errors.append(f"Format invalide en {col_letter}{row_num}: '{cell_value}' (attendu: séquence, temps)")
                            continue
                        
                        sequence = int(float(parts[0].strip()))
                        duration = float(parts[1].strip())
                        
                        if sequence < 1:
                            task_errors.append(f"Séquence invalide en {col_letter}{row_num}: {sequence} (doit être >= 1)")
                            continue
                        
                        if duration <= 0:
                            task_errors.append(f"Durée invalide en {col_letter}{row_num}: {duration} (doit être > 0)")
                            continue
                        
                        # Ajouter la tâche [machine_id, duration] avec la séquence
                        machine_id = j - 3  # Machine 0, 1, 2, etc.
                        job_tasks.append({
                            'sequence': sequence,
                            'machine': machine_id,
                            'duration': duration
                        })
                        
                    except (ValueError, TypeError) as e:
                        task_errors.append(f"Erreur en {col_letter}{row_num}: '{cell_value}' - {str(e)}")
            
            # Vérifier la date due (colonne N, index 13)
            due_date = df.iloc[i, 13]
            due_date_val = 10.0  # valeur par défaut
            
            if pd.notna(due_date) and str(due_date).strip():
                try:
                    due_date_val = float(due_date)
                    if due_date_val < 0:
                        errors.append(f"Date due négative pour '{job_name}' en N{row_num}: {due_date}")
                        due_date_val = 10.0
                except (ValueError, TypeError):
                    errors.append(f"Date due invalide pour '{job_name}' en N{row_num}: '{due_date}' (doit être un nombre)")
                    due_date_val = 10.0
            
            # Ajouter les erreurs de tâches si il y en a
            if task_errors:
                errors.extend([f"Job '{job_name}': {err}" for err in task_errors])
            
            # Trier les tâches par séquence et convertir au format API
            if job_tasks:
                # Trier par séquence
                job_tasks.sort(key=lambda x: x['sequence'])
                
                # Convertir au format API [machine, duration]
                formatted_tasks = [[task['machine'], task['duration']] for task in job_tasks]
                
                job_names.append(job_name)
                jobs_data.append(formatted_tasks)
                due_dates.append(due_date_val)
            elif not task_errors:  # Pas d'erreurs mais pas de tâches non plus
                errors.append(f"Job '{job_name}' (ligne {row_num}): aucune tâche valide trouvée")
        
        # Si il y a des erreurs, les signaler
        if errors:
            error_msg = "Erreurs dans le fichier Excel Jobshop:\n" + "\n".join(f"• {err}" for err in errors[:10])
            if len(errors) > 10:
                error_msg += f"\n... et {len(errors) - 10} autres erreurs"
            raise HTTPException(status_code=400, detail=error_msg)
        
        if not job_names:
            raise HTTPException(
                status_code=400,
                detail="Aucun job valide trouvé. Vérifiez que vous avez rempli les noms de jobs et au moins une tâche par job au format (séquence, temps)."
            )
        
        # Lire les noms des machines depuis les en-têtes (ligne 5, colonnes D-M)
        machine_names = []
        for j in range(3, 13):  # colonnes D à M (index 3-12)
            try:
                header = df.iloc[4, j]  # ligne 5 (index 4)
                if pd.notna(header) and str(header).strip():
                    machine_names.append(str(header).strip())
                else:
                    machine_names.append(f"Machine {j-3}")
            except:
                machine_names.append(f"Machine {j-3}")
        
        # Déterminer le nombre de machines réellement utilisées
        max_machine_id = 0
        for job in jobs_data:
            for task in job:
                max_machine_id = max(max_machine_id, task[0])
        
        # Ajuster les noms de machines
        machine_names = machine_names[:max_machine_id + 1]
        
        return {
            "jobs_data": jobs_data,
            "due_dates": due_dates,
            "job_names": job_names,
            "machine_names": machine_names,
            "unite": unite
        }
        
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=400, detail=f"Erreur lors de la lecture du fichier Excel Jobshop: {str(e)}")

def export_jobshop_data_to_excel(
    jobs_data: List[List[Dict]],  # Format: [[{'sequence': int, 'machine': int, 'duration': float}]]
    due_dates: List[float], 
    job_names: List[str], 
    machine_names: List[str],
    unite: str = "heures"
) -> bytes:
    """
    Exporte les données Jobshop vers un fichier Excel avec le format (séquence, temps)
    
    Args:
        jobs_data: Données des jobs au format Jobshop avec séquence
        due_dates: Dates d'échéance des jobs
        job_names: Noms des jobs
        machine_names: Noms des machines
        unite: Unité de temps
    
    Returns:
        bytes: Contenu du fichier Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Données Jobshop"
    
    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell_alignment = Alignment(horizontal="center", vertical="center")
    
    # Titre principal
    ws['C3'] = "Données d'ordonnancement Jobshop"
    ws['C3'].font = Font(bold=True, size=14)
    ws.merge_cells('C3:M3')
    
    # En-têtes des colonnes
    ws['C5'] = "Job"
    ws['C5'].font = header_font
    ws['C5'].fill = header_fill
    ws['C5'].alignment = cell_alignment
    
    # Noms des machines (colonnes D-M) - TOUJOURS 10 colonnes fixes
    for i in range(10):  # Toujours 10 machines (colonnes D à M)
        col_idx = 4 + i  # Colonne D, E, F, G, H, I, J, K, L, M
        if i < len(machine_names):
            machine_name = machine_names[i]
        else:
            machine_name = f"Machine {i}"
        
        cell = ws.cell(row=5, column=col_idx, value=machine_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = cell_alignment
    
    # Colonne date due - TOUJOURS en colonne N (14)
    due_date_col = 14  # Colonne N fixe
    ws.cell(row=5, column=due_date_col, value="Date due").font = header_font
    ws.cell(row=5, column=due_date_col).fill = header_fill
    ws.cell(row=5, column=due_date_col).alignment = cell_alignment
    
    # Remplir les données des jobs
    for job_idx, (job_name, job_tasks, due_date) in enumerate(zip(job_names, jobs_data, due_dates)):
        row = 6 + job_idx
        
        # Nom du job
        ws.cell(row=row, column=3, value=job_name).alignment = cell_alignment
        
        # Organiser les tâches par machine
        tasks_by_machine = {}
        for task in job_tasks:
            if isinstance(task, dict):
                machine_id = task['machine']
                sequence = task['sequence']
                duration = task['duration']
            else:
                # Fallback pour le format [machine, duration]
                machine_id = task[0]
                duration = task[1]
                sequence = 1  # Séquence par défaut
            
            if machine_id not in tasks_by_machine:
                tasks_by_machine[machine_id] = []
            tasks_by_machine[machine_id].append((sequence, duration))
        
        # Remplir les cellules des machines avec le format séquence, temps
        for machine_id, tasks in tasks_by_machine.items():
            if machine_id < 10:  # Maximum 10 machines (colonnes D à M)
                col_idx = 4 + machine_id  # Colonne D, E, F, G, H, I, J, K, L, M
                
                # Si plusieurs tâches sur la même machine, prendre la première
                if tasks:
                    sequence, duration = tasks[0]
                    cell_value = f"{sequence}, {duration}"
                    ws.cell(row=row, column=col_idx, value=cell_value).alignment = cell_alignment
        
        # Date due
        ws.cell(row=row, column=due_date_col, value=due_date).alignment = cell_alignment
    
    # Unité de temps
    ws['C20'] = unite[0].lower()  # j, h, ou m
    ws['C20'].font = Font(italic=True)
    ws['A20'] = "Unité:"
    ws['A20'].font = Font(italic=True)
    
    # Instructions
    ws['C22'] = "Format des cellules de temps: séquence, durée"
    ws['C22'].font = Font(italic=True, size=10)
    ws['C23'] = "Exemple: 1, 35 = 1ère tâche, durée 35"
    ws['C23'].font = Font(italic=True, size=10)
    
    # Ajuster la largeur des colonnes
    for col in range(3, 15):
        ws.column_dimensions[chr(64 + col)].width = 12
    
    # Sauvegarder
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def parse_flowshop_mm_excel(file_content: bytes) -> Dict:
    """
    Parse un fichier Excel pour l'algorithme Flowshop Machines Multiples
    Format spécifique avec cellules contenant plusieurs durées séparées par des points-virgules
    Exemple: "35; 43.4; 33.5" pour plusieurs machines sur la même étape
    
    Args:
        file_content: Contenu du fichier Excel en bytes
        
    Returns:
        Dict contenant les données formatées pour l'API FlowshopMM
    """
    try:
        # Lire le fichier Excel
        excel_file = io.BytesIO(file_content)
        df = pd.read_excel(excel_file, header=None)
        
        # Vérifier la structure minimale
        if df.shape[0] < 20 or df.shape[1] < 14:
            raise ValueError("Structure de fichier incorrecte")
        
        # Vérifier que c'est bien le format FlowshopMM (cellule C5 doit contenir "Job")
        job_header = df.iloc[4, 2]  # C5 (ligne 5, colonne C)
        if pd.isna(job_header) or str(job_header).strip().lower() != "job":
            raise ValueError("Format non reconnu - cellule C5 doit contenir 'Job'")
        
        # Extraire l'unité de temps (cellule C20)
        unite = "heures"  # valeur par défaut
        try:
            unite_cell = df.iloc[19, 2]  # C20
            if pd.notna(unite_cell):
                unite_str = str(unite_cell).lower().strip()
                if unite_str == 'j':
                    unite = "jours"
                elif unite_str == 'h':
                    unite = "heures"
                elif unite_str == 'm':
                    unite = "minutes"
        except:
            pass
        
        # Extraire les données des jobs
        job_names = []
        jobs_data = []
        due_dates = []
        errors = []
        machines_per_stage = []
        
        # Parcourir les lignes 6-15 (index 5-14) pour les jobs
        for i in range(5, 15):
            row_num = i + 1  # Numéro de ligne Excel (1-indexé)
            
            # Nom du job (colonne C)
            job_name = df.iloc[i, 2]
            if pd.isna(job_name) or not str(job_name).strip():
                continue  # Ligne vide, on passe
            
            job_name = str(job_name).strip()
            
            # Extraire les durées du job (colonnes D-M, index 3-12)
            job_stages = []
            stage_errors = []
            
            for j in range(3, 13):  # colonnes D à M (10 étapes max)
                col_letter = chr(68 + j - 3)  # D, E, F, G, H, I, J, K, L, M
                cell_value = df.iloc[i, j]
                
                if pd.notna(cell_value) and str(cell_value).strip():
                    try:
                        # Parser le format avec plusieurs durées séparées par des points-virgules
                        cell_str = str(cell_value).strip()
                        
                        # Séparer par point-virgule
                        duration_parts = [part.strip() for part in cell_str.split(';')]
                        
                        stage_alternatives = []
                        for alt_index, duration_str in enumerate(duration_parts):
                            if duration_str:  # Ignorer les parties vides
                                try:
                                    duration = float(duration_str)
                                    if duration < 0:
                                        stage_errors.append(f"Durée négative en {col_letter}{row_num}: {duration}")
                                        continue
                                    
                                    # Machine ID : étape (base 1) * 10 + (alternative + 1)
                                    machine_id = (j - 2) * 10 + (alt_index + 1)  # j-2 car colonne D = étape 1
                                    stage_alternatives.append([machine_id, duration])
                                    
                                except (ValueError, TypeError):
                                    stage_errors.append(f"Valeur invalide en {col_letter}{row_num}: '{duration_str}' (doit être un nombre)")
                        
                        if stage_alternatives:
                            job_stages.append(stage_alternatives)
                            
                            # Mettre à jour le nombre max de machines par étape
                            stage_index = j - 3
                            while len(machines_per_stage) <= stage_index:
                                machines_per_stage.append(1)
                            machines_per_stage[stage_index] = max(machines_per_stage[stage_index], len(stage_alternatives))
                        
                    except Exception as e:
                        stage_errors.append(f"Erreur en {col_letter}{row_num}: '{cell_value}' - {str(e)}")
            
            # Vérifier la date due (colonne N, index 13)
            due_date = df.iloc[i, 13]
            due_date_val = 10.0  # valeur par défaut
            
            if pd.notna(due_date) and str(due_date).strip():
                try:
                    due_date_val = float(due_date)
                    if due_date_val < 0:
                        errors.append(f"Date due négative pour '{job_name}' en N{row_num}: {due_date}")
                        due_date_val = 10.0
                except (ValueError, TypeError):
                    errors.append(f"Date due invalide pour '{job_name}' en N{row_num}: '{due_date}' (doit être un nombre)")
                    due_date_val = 10.0
            
            # Ajouter les erreurs d'étapes si il y en a
            if stage_errors:
                errors.extend([f"Job '{job_name}': {err}" for err in stage_errors])
            
            # Ajouter le job seulement s'il a au moins une étape valide
            if job_stages:
                job_names.append(job_name)
                jobs_data.append(job_stages)
                due_dates.append(due_date_val)
            elif not stage_errors:  # Pas d'erreurs mais pas d'étapes non plus
                errors.append(f"Job '{job_name}' (ligne {row_num}): aucune étape valide trouvée")
        
        # Si il y a des erreurs, les signaler
        if errors:
            error_msg = "Erreurs dans le fichier Excel FlowshopMM:\n" + "\n".join(f"• {err}" for err in errors[:10])
            if len(errors) > 10:
                error_msg += f"\n... et {len(errors) - 10} autres erreurs"
            raise HTTPException(status_code=400, detail=error_msg)
        
        if not job_names:
            raise HTTPException(
                status_code=400,
                detail="Aucun job valide trouvé. Vérifiez que vous avez rempli les noms de jobs et au moins une durée par job (format: 35 ou 35; 43.4; 33.5)."
            )
        
        # Lire les noms des machines depuis les en-têtes (ligne 5, colonnes D-M)
        stage_names = []
        for j in range(3, 13):  # colonnes D à M (index 3-12)
            try:
                header = df.iloc[4, j]  # ligne 5 (index 4)
                if pd.notna(header) and str(header).strip():
                    stage_names.append(str(header).strip())
                else:
                    stage_names.append(f"Étape {j-2}")
            except:
                stage_names.append(f"Étape {j-2}")
        
        # Déterminer le nombre d'étapes réellement utilisées
        max_stage_count = len(jobs_data[0]) if jobs_data else 0
        for job in jobs_data:
            max_stage_count = max(max_stage_count, len(job))
        
        # Ajuster les noms d'étapes et machines_per_stage
        stage_names = stage_names[:max_stage_count]
        machines_per_stage = machines_per_stage[:max_stage_count]
        
        # Générer les priorités des machines (ordre d'apparition)
        machine_priorities = {}
        for job in jobs_data:
            for stage in job:
                for alt_index, (machine_id, duration) in enumerate(stage):
                    if machine_id not in machine_priorities:
                        machine_priorities[machine_id] = alt_index + 1
        
        return {
            "jobs_data": jobs_data,
            "due_dates": due_dates,
            "job_names": job_names,
            "stage_names": stage_names,
            "machines_per_stage": machines_per_stage,
            "machine_priorities": machine_priorities,
            "unite": unite
        }
        
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=400, detail=f"Erreur lors de la lecture du fichier Excel FlowshopMM: {str(e)}")

def export_flowshop_mm_data_to_excel(
    jobs_data: List[List[List[List[float]]]],  # Format FlowshopMM: job -> stage -> alternatives -> [machine_id, duration]
    due_dates: List[float], 
    job_names: List[str], 
    stage_names: List[str],
    machines_per_stage: List[int],
    unite: str = "heures"
) -> bytes:
    """
    Exporte les données FlowshopMM vers un fichier Excel avec le format séparé par points-virgules
    
    Args:
        jobs_data: Données des jobs au format FlowshopMM
        due_dates: Dates d'échéance des jobs
        job_names: Noms des jobs
        stage_names: Noms des étapes
        machines_per_stage: Nombre de machines par étape
        unite: Unité de temps
    
    Returns:
        bytes: Contenu du fichier Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Données FlowshopMM"
    
    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell_alignment = Alignment(horizontal="center", vertical="center")
    
    # Titre principal
    ws['C3'] = "Données d'ordonnancement Flowshop Machines Multiples"
    ws['C3'].font = Font(bold=True, size=14)
    ws.merge_cells('C3:M3')
    
    # En-têtes des colonnes
    ws['C5'] = "Job"
    ws['C5'].font = header_font
    ws['C5'].fill = header_fill
    ws['C5'].alignment = cell_alignment
    
    # Noms des étapes (colonnes D-M) - TOUJOURS 10 colonnes fixes
    for i in range(10):  # Toujours 10 étapes (colonnes D à M)
        col_idx = 4 + i  # Colonne D, E, F, G, H, I, J, K, L, M
        if i < len(stage_names):
            stage_name = stage_names[i]
        else:
            stage_name = f"Étape {i + 1}"
        
        cell = ws.cell(row=5, column=col_idx, value=stage_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = cell_alignment
    
    # Colonne date due - TOUJOURS en colonne N (14)
    due_date_col = 14  # Colonne N fixe
    ws.cell(row=5, column=due_date_col, value="Date due").font = header_font
    ws.cell(row=5, column=due_date_col).fill = header_fill
    ws.cell(row=5, column=due_date_col).alignment = cell_alignment
    
    # Remplir les données des jobs
    for job_idx, (job_name, job_stages, due_date) in enumerate(zip(job_names, jobs_data, due_dates)):
        row = 6 + job_idx
        
        # Nom du job
        ws.cell(row=row, column=3, value=job_name).alignment = cell_alignment
        
        # Remplir les cellules des étapes avec le format séparé par points-virgules
        for stage_idx, stage_alternatives in enumerate(job_stages):
            if stage_idx < 10:  # Maximum 10 étapes (colonnes D à M)
                col_idx = 4 + stage_idx  # Colonne D, E, F, G, H, I, J, K, L, M
                
                # Extraire seulement les durées et les joindre avec des points-virgules
                durations = [str(alt[1]) for alt in stage_alternatives]  # alt[1] = duration
                cell_value = "; ".join(durations)
                
                ws.cell(row=row, column=col_idx, value=cell_value).alignment = cell_alignment
        
        # Date due
        ws.cell(row=row, column=due_date_col, value=due_date).alignment = cell_alignment
    
    # Unité de temps
    ws['C20'] = unite[0].lower()  # j, h, ou m
    ws['C20'].font = Font(italic=True)
    ws['A20'] = "Unité:"
    ws['A20'].font = Font(italic=True)
    
    # Instructions
    ws['C22'] = "Format des cellules: une ou plusieurs durées séparées par des points-virgules"
    ws['C22'].font = Font(italic=True, size=10)
    ws['C23'] = "Exemple: 35 (une machine) ou 35; 43.4; 33.5 (plusieurs machines)"
    ws['C23'].font = Font(italic=True, size=10)
    
    # Ajuster la largeur des colonnes
    for col in range(3, 15):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Sauvegarder
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def export_manual_data_to_excel(
    jobs_data: List[List[float]], 
    due_dates: List[float], 
    job_names: List[str], 
    machine_names: List[str],
    unite: str = "heures"
) -> bytes:
    """
    Exporte les données saisies manuellement vers un fichier Excel au format matriciel FIXE 12x12
    Structure exacte du template (TOUJOURS 12 colonnes et 12 lignes) :
    - C5: "JOB"
    - C6-C16: Noms des jobs (max 11 jobs)
    - D5-M5: Noms des machines (max 10 machines) 
    - N5: "Due Date" (TOUJOURS colonne N = 14)
    - N6-N16: Dates d'échéance 
    - D6-M16: Matrice des temps de traitement
    
    Args:
        jobs_data: Données des jobs [[durée_machine_0, durée_machine_1, ...], ...]
        due_dates: Dates d'échéance des jobs
        job_names: Noms des jobs
        machine_names: Noms des machines
        unite: Unité de temps
        
    Returns:
        bytes: Contenu du fichier Excel
    """
    # Validation des données d'entrée
    if not jobs_data or not due_dates or not job_names or not machine_names:
        raise ValueError("Toutes les données (jobs_data, due_dates, job_names, machine_names) sont requises")
    
    # Vérifier que toutes les listes ont la même longueur
    num_jobs = len(job_names)
    if len(jobs_data) != num_jobs:
        raise ValueError(f"Le nombre de jobs dans jobs_data ({len(jobs_data)}) ne correspond pas au nombre de noms de jobs ({num_jobs})")
    
    if len(due_dates) != num_jobs:
        raise ValueError(f"Le nombre de dates d'échéance ({len(due_dates)}) ne correspond pas au nombre de jobs ({num_jobs})")
    
    # Debug : afficher les données reçues
    print(f"DEBUG - jobs_data: {jobs_data}")
    print(f"DEBUG - due_dates: {due_dates}")
    print(f"DEBUG - job_names: {job_names}")
    print(f"DEBUG - machine_names: {machine_names}")
    
    # Normaliser les données des jobs pour s'assurer qu'elles ont toutes la bonne longueur
    normalized_jobs_data = []
    for job_idx, job in enumerate(jobs_data):
        try:
            if isinstance(job, list):
                # Étendre ou tronquer le job pour qu'il ait exactement le nombre de machines
                normalized_job = []
                for i in range(len(machine_names)):
                    if i < len(job):
                        duration = float(job[i]) if job[i] is not None else 0.0
                    else:
                        duration = 0.0
                    normalized_job.append(duration)
                normalized_jobs_data.append(normalized_job)
                print(f"DEBUG - Job {job_idx} ({job_names[job_idx]}): {job} -> {normalized_job}")
            else:
                # Si ce n'est pas une liste, créer une liste de zéros
                normalized_jobs_data.append([0.0] * len(machine_names))
                print(f"DEBUG - Job {job_idx} ({job_names[job_idx]}): Not a list, created zeros")
        except (ValueError, TypeError) as e:
            raise ValueError(f"Erreur dans les données du job '{job_names[job_idx]}': {str(e)}")
    
    # Créer un BytesIO pour le fichier Excel
    output = io.BytesIO()
    
    # Créer un workbook avec openpyxl pour un contrôle précis de la structure
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Données"
    
    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # STRUCTURE FIXE 12x12 - TOUJOURS utiliser ces positions exactes
    
    # C5: "JOB" (coin supérieur gauche)
    ws['C5'] = "JOB"
    ws['C5'].font = header_font
    ws['C5'].fill = header_fill
    ws['C5'].alignment = Alignment(horizontal="center")
    ws['C5'].border = border
    
    # C6-C16: Noms des jobs (colonne C = 3, lignes 6 à 16 = max 11 jobs)
    for i in range(11):  # TOUJOURS 11 lignes (C6 à C16)
        if i < len(job_names):
            job_name = job_names[i]
        else:
            job_name = ""  # Cellule vide si pas assez de jobs
        
        cell = ws.cell(row=6+i, column=3, value=job_name)  # Colonne C = 3
        cell.border = border
    
    # D5-M5: Noms des machines (ligne 5, colonnes D=4 à M=13 = max 10 machines)
    for i in range(10):  # TOUJOURS 10 colonnes (D à M)
        if i < len(machine_names):
            machine_name = machine_names[i]
        else:
            machine_name = ""  # Cellule vide si pas assez de machines
            
        cell = ws.cell(row=5, column=4+i, value=machine_name)  # Colonnes D=4 à M=13
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    # N5: "Due Date" (TOUJOURS colonne N = 14)
    ws.cell(row=5, column=14, value="Due Date")  # Colonne N = 14
    ws.cell(row=5, column=14).font = header_font
    ws.cell(row=5, column=14).fill = header_fill
    ws.cell(row=5, column=14).alignment = Alignment(horizontal="center")
    ws.cell(row=5, column=14).border = border
    
    # N6-N16: Dates d'échéance (colonne N=14, lignes 6 à 16)
    for i in range(11):  # TOUJOURS 11 lignes (N6 à N16)
        if i < len(due_dates):
            due_date = due_dates[i]
        else:
            due_date = ""  # Cellule vide si pas assez de dates
            
        cell = ws.cell(row=6+i, column=14, value=due_date)  # Colonne N = 14
        cell.border = border
    
    # D6-M16: Matrice des temps de traitement (colonnes D=4 à M=13, lignes 6 à 16)
    for job_idx in range(11):  # TOUJOURS 11 lignes de jobs
        for machine_idx in range(10):  # TOUJOURS 10 colonnes de machines
            if job_idx < len(normalized_jobs_data) and machine_idx < len(normalized_jobs_data[job_idx]):
                duration = normalized_jobs_data[job_idx][machine_idx]
                print(f"DEBUG - Writing job {job_idx}, machine {machine_idx}: duration = {duration} (type: {type(duration)})")
            else:
                duration = ""  # Cellule vide si pas de données
                print(f"DEBUG - No data for job {job_idx}, machine {machine_idx}: setting empty")
                
            cell = ws.cell(row=6+job_idx, column=4+machine_idx, value=duration)
            cell.border = border
            print(f"DEBUG - Cell ({6+job_idx}, {4+machine_idx}) = {cell.value}")
    
    # Ajuster la largeur des colonnes (C à N)
    for col in range(3, 15):  # De C=3 à N=14
        column_letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[column_letter].width = 12
    
    # Ajouter un onglet d'instructions
    instructions_ws = wb.create_sheet("Instructions")
    instructions_ws['A1'] = "DONNÉES EXPORTÉES - Format Matriciel 12x12"
    instructions_ws['A1'].font = Font(bold=True, size=14)
    
    instructions = [
        "",
        "Structure FIXE du fichier (12 colonnes x 12 lignes) :",
        "- C5: 'JOB' (coin supérieur gauche)",
        "- C6-C16: Noms des jobs (max 11 jobs)",
        "- D5-M5: Noms des machines (max 10 machines)",
        "- N5: 'Due Date' (TOUJOURS colonne N)",
        "- N6-N16: Dates d'échéance",
        "- D6-M16: Matrice des temps de traitement",
        "",
        "Paramètres actuels :",
        f"- Unité de temps: {unite}",
        f"- Nombre de machines utilisées: {len(machine_names)}/10",
        f"- Nombre de jobs utilisés: {num_jobs}/11",
        "",
        "IMPORTANT :",
        "- La structure 12x12 est FIXE pour compatibilité d'import",
        "- Les cellules vides sont normales si moins de données",
        "- La colonne 'Due Date' est TOUJOURS en colonne N",
        "",
        "Ce fichier peut être modifié et réimporté dans l'application."
    ]
    
    for i, instruction in enumerate(instructions):
        instructions_ws.cell(row=2+i, column=1, value=instruction)
    
    # C19: "Unité de temps"
    ws.cell(row=19, column=3, value="Unité de temps")
    
    # C20: Unité de temps (j/h/m selon l'interface)
    unit_mapping = {
        "jours": "j",
        "heures": "h", 
        "minutes": "m",
        "jour": "j",
        "heure": "h",
        "minute": "m"
    }
    
    # Conversion de l'unité reçue en abréviation
    unit_abbrev = unit_mapping.get(unite.lower(), unite.lower())
    ws.cell(row=20, column=3, value=unit_abbrev)
    
    print(f"DEBUG - Unité de temps ajoutée: '{unite}' -> '{unit_abbrev}' en C20")

    # Sauvegarder dans le BytesIO
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ===== FONCTIONS SPÉCIFIQUES POUR LIGNE D'ASSEMBLAGE =====

async def parse_ligne_assemblage_excel(file) -> Dict:
    """
    Parse un fichier Excel pour les algorithmes de ligne d'assemblage.
    Format attendu :
    - C6: "Tâche" (header)
    - C7+: Noms des tâches
    - D6: "Durée" (header)
    - D7+: Durées des tâches
    - E6: "Prédécesseur" (header)  
    - E7+: Prédécesseurs des tâches
    
    Args:
        file: Fichier Excel uploadé
        
    Returns:
        Dict contenant les données formatées pour l'API ligne d'assemblage
    """
    try:
        # Lire le contenu du fichier
        file_content = await file.read()
        
        # Lire le fichier Excel sans en-tête automatique
        excel_file = io.BytesIO(file_content)
        df = pd.read_excel(excel_file, header=None)
        
        # Vérifier la structure minimale
        if df.shape[0] < 10 or df.shape[1] < 6:
            raise ValueError("Structure de fichier incorrecte")
        
        # Vérifier les headers en ligne 6 (index 5)
        task_header = df.iloc[5, 2]  # C6
        duration_header = df.iloc[5, 3]  # D6
        predecessor_header = df.iloc[5, 4]  # E6
        
        if (pd.isna(task_header) or str(task_header).strip().lower() != "tâche" and 
            str(task_header).strip().lower() != "tache"):
            raise ValueError("Format non reconnu - cellule C6 doit contenir 'Tâche'")
        
        # Extraire les données des tâches
        tasks_data = []
        cycle_time = 70.0  # valeur par défaut
        unite = "minutes"  # valeur par défaut
        errors = []
        
        # Parcourir les lignes à partir de la ligne 7 (index 6)
        task_id = 1
        for i in range(6, min(df.shape[0], 50)):  # Limiter à 50 tâches max
            row_num = i + 1  # Numéro de ligne Excel (1-indexé)
            
            # Nom de la tâche (colonne C)
            task_name = df.iloc[i, 2]
            if pd.isna(task_name) or not str(task_name).strip():
                continue  # Ligne vide, on passe
            
            task_name = str(task_name).strip()
            
            # Durée de la tâche (colonne D)
            duration = df.iloc[i, 3]
            if pd.isna(duration) or not str(duration).strip():
                errors.append(f"Tâche '{task_name}' (ligne {row_num}): durée manquante")
                continue
            
            try:
                duration_val = float(duration)
                if duration_val <= 0:
                    errors.append(f"Tâche '{task_name}' (ligne {row_num}): durée doit être positive")
                    continue
            except (ValueError, TypeError):
                errors.append(f"Tâche '{task_name}' (ligne {row_num}): durée invalide '{duration}'")
                continue
            
            # Prédécesseurs (colonne E)
            predecessors = df.iloc[i, 4]
            predecessors_val = None
            
            if pd.notna(predecessors) and str(predecessors).strip():
                predecessors_str = str(predecessors).strip()
                if predecessors_str and predecessors_str != "0":
                    # Parser les prédécesseurs (format: "1,2,3" ou "1")
                    try:
                        pred_list = [int(p.strip()) for p in predecessors_str.split(',') if p.strip()]
                        if len(pred_list) == 1:
                            predecessors_val = pred_list[0]
                        elif len(pred_list) > 1:
                            predecessors_val = pred_list
                    except (ValueError, TypeError):
                        errors.append(f"Tâche '{task_name}' (ligne {row_num}): prédécesseurs invalides '{predecessors_str}'")
            
            # Ajouter la tâche
            tasks_data.append({
                "id": task_id,
                "name": task_name,
                "predecessors": predecessors_val,
                "duration": duration_val
            })
            
            task_id += 1
        
        # Chercher le temps de cycle et l'unité aux positions spécifiques
        # H6 = "Unité de temps", H7 = valeur unité
        # H9 = "Temps de cycle", H10 = valeur temps de cycle
        try:
            # Vérifier H6 pour "Unité de temps"
            if df.shape[0] > 5 and df.shape[1] > 7:  # H6 = ligne 6, colonne H (index 7)
                h6_value = df.iloc[5, 7]  # H6
                if pd.notna(h6_value) and "unité" in str(h6_value).lower():
                    # Lire la valeur de l'unité en H7
                    h7_value = df.iloc[6, 7]  # H7
                    if pd.notna(h7_value):
                        unite_val = str(h7_value).strip().lower()
                        if unite_val in ['j', 'h', 'm']:
                            if unite_val == 'j':
                                unite = "jours"
                            elif unite_val == 'h':
                                unite = "heures"
                            elif unite_val == 'm':
                                unite = "minutes"
            
            # Vérifier H9 pour "Temps de cycle"
            if df.shape[0] > 9 and df.shape[1] > 7:  # H9 = ligne 9, colonne H (index 7)
                h9_value = df.iloc[8, 7]  # H9
                if pd.notna(h9_value) and "cycle" in str(h9_value).lower():
                    # Lire la valeur du temps de cycle en H10
                    h10_value = df.iloc[9, 7]  # H10
                    if pd.notna(h10_value):
                        cycle_time = float(h10_value)
        except:
            pass  # Garder les valeurs par défaut si erreur
        
        # Si il y a des erreurs, les signaler
        if errors:
            error_msg = "Erreurs dans le fichier Excel:\n" + "\n".join(f"• {err}" for err in errors[:10])
            if len(errors) > 10:
                error_msg += f"\n... et {len(errors) - 10} autres erreurs"
            raise HTTPException(status_code=400, detail=error_msg)
        
        if not tasks_data:
            raise HTTPException(
                status_code=400,
                detail="Aucune tâche valide trouvée. Vérifiez que vous avez rempli les noms, durées et prédécesseurs."
            )
        
        return {
            "tasks_data": tasks_data,
            "cycle_time": cycle_time,
            "unite": unite
        }
        
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=400, detail=f"Erreur lors de la lecture du fichier Excel: {str(e)}")


def export_ligne_assemblage_to_excel(
    tasks_data: List[dict],
    cycle_time: float,
    unite: str = "minutes",
    algorithm_name: str = "Ligne d'assemblage"
) -> bytes:
    """
    Exporte les données de ligne d'assemblage vers Excel.
    Format de sortie :
    - B7+: ID des tâches (1, 2, 3...)
    - C6: "Tâche", C7+: Noms des tâches
    - D6: "Durée", D7+: Durées des tâches
    - E6: "Prédécesseur", E7+: Prédécesseurs des tâches
    
    Args:
        tasks_data: Liste des tâches avec task_id, name, duration, predecessors
        cycle_time: Temps de cycle
        unite: Unité de temps
        algorithm_name: Nom de l'algorithme
        
    Returns:
        bytes: Contenu du fichier Excel
    """
    try:
        from fastapi.responses import StreamingResponse
        
        # Validation des données d'entrée
        if not tasks_data:
            raise ValueError("Aucune donnée de tâche fournie")
        
        # Créer un BytesIO pour le fichier Excel
        output = io.BytesIO()
        
        # Créer un workbook avec openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Export_{algorithm_name}"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers en ligne 6
        # B6: "ID" (pour aider l'utilisateur)
        ws['B6'] = "ID"
        ws['B6'].font = header_font
        ws['B6'].fill = header_fill
        ws['B6'].alignment = Alignment(horizontal="center")
        ws['B6'].border = border
        
        # C6: "Tâche"
        ws['C6'] = "Tâche"
        ws['C6'].font = header_font
        ws['C6'].fill = header_fill
        ws['C6'].alignment = Alignment(horizontal="center")
        ws['C6'].border = border
        
        # D6: "Durée"
        ws['D6'] = "Durée"
        ws['D6'].font = header_font
        ws['D6'].fill = header_fill
        ws['D6'].alignment = Alignment(horizontal="center")
        ws['D6'].border = border
        
        # E6: "Prédécesseur"
        ws['E6'] = "Prédécesseur"
        ws['E6'].font = header_font
        ws['E6'].fill = header_fill
        ws['E6'].alignment = Alignment(horizontal="center")
        ws['E6'].border = border
        
        # Données à partir de la ligne 7
        for i, task in enumerate(tasks_data):
            row = 7 + i
            
            # B7+: ID des tâches (task_id ou index+1)
            task_id = task.get("task_id", i + 1)
            ws.cell(row=row, column=2, value=task_id).border = border
            
            # C7+: Noms des tâches
            ws.cell(row=row, column=3, value=task.get("name", f"Tâche {task_id}")).border = border
            
            # D7+: Durées des tâches
            ws.cell(row=row, column=4, value=task.get("duration", 0)).border = border
            
            # E7+: Prédécesseurs des tâches
            predecessors = task.get("predecessors")
            if predecessors is None:
                pred_str = ""
            elif isinstance(predecessors, list):
                pred_str = ",".join(map(str, predecessors))
            else:
                pred_str = str(predecessors)
            
            ws.cell(row=row, column=5, value=pred_str).border = border
        
        # Informations supplémentaires - Placer dans la colonne H aux positions demandées
        # H6: "Unité de temps"
        ws['H6'] = "Unité de temps"
        ws['H6'].font = Font(bold=True)
        
        # H7: Valeur de l'unité (j/h/m)
        # Convertir l'unité au format court
        unite_short = unite
        if unite.lower() == "minutes":
            unite_short = "m"
        elif unite.lower() == "heures":
            unite_short = "h"
        elif unite.lower() == "jours":
            unite_short = "j"
        ws['H7'] = unite_short
        
        # H9: "Temps de cycle"
        ws['H9'] = "Temps de cycle"
        ws['H9'].font = Font(bold=True)
        
        # H10: Valeur du temps de cycle
        ws['H10'] = cycle_time
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        
        # Ajouter un onglet d'instructions
        instructions_ws = wb.create_sheet("Instructions")
        instructions_ws['A1'] = f"EXPORT {algorithm_name.upper()} - LIGNE D'ASSEMBLAGE"
        instructions_ws['A1'].font = Font(bold=True, size=14)
        
        instructions = [
            "",
            "Structure du fichier :",
            "- Colonne B: ID des tâches (pour référence, non importé)",
            "- Colonne C: Noms des tâches",
            "- Colonne D: Durées des tâches",
            "- Colonne E: Prédécesseurs (vide si aucun, sinon IDs séparés par virgules)",
            "- Colonne H6: Unité de temps, H7: Valeur unité (j/h/m)",
            "- Colonne H9: Temps de cycle, H10: Valeur temps de cycle",
            "",
            "Format d'import attendu :",
            "- Headers en ligne 6 (C6=Tâche, D6=Durée, E6=Prédécesseur)",
            "- Données à partir de la ligne 7",
            "- Paramètres en colonne H aux lignes spécifiées",
            "",
            f"Nombre de tâches exportées: {len(tasks_data)}",
            "",
            "Ce fichier peut être modifié et réimporté dans l'application."
        ]
        
        for i, instruction in enumerate(instructions):
            instructions_ws.cell(row=2+i, column=1, value=instruction)
        
        # Sauvegarder dans le BytesIO
        wb.save(output)
        output.seek(0)
        
        # Retourner une réponse de streaming
        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=Export_{algorithm_name}_LigneAssemblage.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'export: {str(e)}")

async def parse_precedence_excel(file) -> Dict:
    """
    Parse un fichier Excel pour l'algorithme de précédences.
    Format attendu identique à ligne d'assemblage mais sans cycle_time obligatoire.
    - C6: "Tâche" (header)
    - C7+: Noms des tâches
    - D6: "Durée" (header)
    - D7+: Durées des tâches
    - E6: "Prédécesseur" (header)  
    - E7+: Prédécesseurs des tâches
    - H6: "Unité de temps", H7: Valeur unité (j/h/m) - optionnel
    - H9: "Temps de cycle", H10: Valeur (ignoré pour précédences)
    
    Args:
        file: Fichier Excel uploadé
        
    Returns:
        Dict contenant les données formatées pour l'API précédences
    """
    try:
        # Lire le contenu du fichier
        file_content = await file.read()
        
        # Lire le fichier Excel sans en-tête automatique
        excel_file = io.BytesIO(file_content)
        df = pd.read_excel(excel_file, header=None)
        
        # Vérifier la structure minimale
        if df.shape[0] < 10 or df.shape[1] < 6:
            raise ValueError("Structure de fichier incorrecte")
        
        # Vérifier les headers en ligne 6 (index 5)
        task_header = df.iloc[5, 2]  # C6
        duration_header = df.iloc[5, 3]  # D6
        predecessor_header = df.iloc[5, 4]  # E6
        
        if (pd.isna(task_header) or str(task_header).strip().lower() != "tâche" and 
            str(task_header).strip().lower() != "tache"):
            raise ValueError("Format non reconnu - cellule C6 doit contenir 'Tâche'")
        
        # Extraire les données des tâches
        tasks_data = []
        unite = "minutes"  # valeur par défaut
        errors = []
        
        # Parcourir les lignes à partir de la ligne 7 (index 6)
        task_id = 1
        for i in range(6, min(df.shape[0], 50)):  # Limiter à 50 tâches max
            row_num = i + 1  # Numéro de ligne Excel (1-indexé)
            
            # Nom de la tâche (colonne C)
            task_name = df.iloc[i, 2]
            if pd.isna(task_name) or not str(task_name).strip():
                continue  # Ligne vide, on passe
            
            task_name = str(task_name).strip()
            
            # Durée de la tâche (colonne D)
            duration = df.iloc[i, 3]
            if pd.isna(duration) or not str(duration).strip():
                errors.append(f"Tâche '{task_name}' (ligne {row_num}): durée manquante")
                continue
            
            try:
                duration_val = float(duration)
                if duration_val <= 0:
                    errors.append(f"Tâche '{task_name}' (ligne {row_num}): durée doit être positive")
                    continue
            except (ValueError, TypeError):
                errors.append(f"Tâche '{task_name}' (ligne {row_num}): durée invalide '{duration}'")
                continue
            
            # Prédécesseurs (colonne E)
            predecessors = df.iloc[i, 4]
            predecessors_val = None
            
            if pd.notna(predecessors) and str(predecessors).strip():
                predecessors_str = str(predecessors).strip()
                if predecessors_str and predecessors_str != "0":
                    # Parser les prédécesseurs (format: "1,2,3" ou "1")
                    try:
                        pred_list = [int(p.strip()) for p in predecessors_str.split(',') if p.strip()]
                        if len(pred_list) == 1:
                            predecessors_val = pred_list[0]
                        elif len(pred_list) > 1:
                            predecessors_val = pred_list
                    except (ValueError, TypeError):
                        errors.append(f"Tâche '{task_name}' (ligne {row_num}): prédécesseurs invalides '{predecessors_str}'")
            
            # Ajouter la tâche
            tasks_data.append({
                "id": task_id,
                "name": task_name,
                "predecessors": predecessors_val,
                "duration": duration_val
            })
            
            task_id += 1
        
        # Chercher l'unité aux positions spécifiques (optionnel pour précédences)
        try:
            # Vérifier H6 pour "Unité de temps"
            if df.shape[0] > 5 and df.shape[1] > 7:  # H6 = ligne 6, colonne H (index 7)
                h6_value = df.iloc[5, 7]  # H6
                if pd.notna(h6_value) and "unité" in str(h6_value).lower():
                    # Lire la valeur de l'unité en H7
                    h7_value = df.iloc[6, 7]  # H7
                    if pd.notna(h7_value):
                        unite_val = str(h7_value).strip().lower()
                        if unite_val in ['j', 'h', 'm']:
                            if unite_val == 'j':
                                unite = "jours"
                            elif unite_val == 'h':
                                unite = "heures"
                            elif unite_val == 'm':
                                unite = "minutes"
        except:
            pass  # Garder la valeur par défaut si erreur
        
        # Si il y a des erreurs, les signaler
        if errors:
            error_msg = "Erreurs dans le fichier Excel:\n" + "\n".join(f"• {err}" for err in errors[:10])
            if len(errors) > 10:
                error_msg += f"\n... et {len(errors) - 10} autres erreurs"
            raise HTTPException(status_code=400, detail=error_msg)
        
        if not tasks_data:
            raise HTTPException(
                status_code=400,
                detail="Aucune tâche valide trouvée. Vérifiez que vous avez rempli les noms, durées et prédécesseurs."
            )
        
        return {
            "tasks_data": tasks_data,
            "unite": unite
        }
        
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=400, detail=f"Erreur lors de la lecture du fichier Excel: {str(e)}")


def export_precedence_to_excel(
    tasks_data: List[dict],
    unite: str = "minutes",
    algorithm_name: str = "Précédences"
) -> bytes:
    """
    Exporte les données de précédences vers Excel.
    Format de sortie identique à ligne d'assemblage mais avec cycle_time par défaut (non utilisé).
    - B7+: ID des tâches (1, 2, 3...)
    - C6: "Tâche", C7+: Noms des tâches
    - D6: "Durée", D7+: Durées des tâches
    - E6: "Prédécesseur", E7+: Prédécesseurs des tâches
    - H6: "Unité de temps", H7: Valeur unité (j/h/m)
    - H9: "Temps de cycle", H10: 70 (valeur par défaut, non utilisée)
    
    Args:
        tasks_data: Liste des tâches avec task_id, name, duration, predecessors
        unite: Unité de temps
        algorithm_name: Nom de l'algorithme
        
    Returns:
        bytes: Contenu du fichier Excel
    """
    try:
        from fastapi.responses import StreamingResponse
        
        # Validation des données d'entrée
        if not tasks_data:
            raise ValueError("Aucune donnée de tâche fournie")
        
        # Créer un BytesIO pour le fichier Excel
        output = io.BytesIO()
        
        # Créer un workbook avec openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Export_{algorithm_name}"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers en ligne 6
        # B6: "ID" (pour aider l'utilisateur)
        ws['B6'] = "ID"
        ws['B6'].font = header_font
        ws['B6'].fill = header_fill
        ws['B6'].alignment = Alignment(horizontal="center")
        ws['B6'].border = border
        
        # C6: "Tâche"
        ws['C6'] = "Tâche"
        ws['C6'].font = header_font
        ws['C6'].fill = header_fill
        ws['C6'].alignment = Alignment(horizontal="center")
        ws['C6'].border = border
        
        # D6: "Durée"
        ws['D6'] = "Durée"
        ws['D6'].font = header_font
        ws['D6'].fill = header_fill
        ws['D6'].alignment = Alignment(horizontal="center")
        ws['D6'].border = border
        
        # E6: "Prédécesseur"
        ws['E6'] = "Prédécesseur"
        ws['E6'].font = header_font
        ws['E6'].fill = header_fill
        ws['E6'].alignment = Alignment(horizontal="center")
        ws['E6'].border = border
        
        # Données à partir de la ligne 7
        for i, task in enumerate(tasks_data):
            row = 7 + i
            
            # B7+: ID des tâches (task_id ou index+1)
            task_id = task.get("task_id", i + 1)
            ws.cell(row=row, column=2, value=task_id).border = border
            
            # C7+: Noms des tâches
            ws.cell(row=row, column=3, value=task.get("name", f"Tâche {task_id}")).border = border
            
            # D7+: Durées des tâches
            ws.cell(row=row, column=4, value=task.get("duration", 0)).border = border
            
            # E7+: Prédécesseurs des tâches
            predecessors = task.get("predecessors")
            if predecessors is None:
                pred_str = ""
            elif isinstance(predecessors, list):
                pred_str = ",".join(map(str, predecessors))
            else:
                pred_str = str(predecessors)
            
            ws.cell(row=row, column=5, value=pred_str).border = border
        
        # Informations supplémentaires - Placer dans la colonne H aux positions demandées
        # H6: "Unité de temps"
        ws['H6'] = "Unité de temps"
        ws['H6'].font = Font(bold=True)
        
        # H7: Valeur de l'unité (j/h/m)
        # Convertir l'unité au format court
        unite_short = unite
        if unite.lower() == "minutes":
            unite_short = "m"
        elif unite.lower() == "heures":
            unite_short = "h"
        elif unite.lower() == "jours":
            unite_short = "j"
        ws['H7'] = unite_short
        
        # H9: "Temps de cycle" (pour compatibilité, mais non utilisé par précédences)
        ws['H9'] = "Temps de cycle"
        ws['H9'].font = Font(bold=True)
        
        # H10: Valeur du temps de cycle par défaut (non utilisée par précédences)
        ws['H10'] = 70
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        
        # Ajouter un onglet d'instructions
        instructions_ws = wb.create_sheet("Instructions")
        instructions_ws['A1'] = f"EXPORT {algorithm_name.upper()} - DIAGRAMME DE PRÉCÉDENCE"
        instructions_ws['A1'].font = Font(bold=True, size=14)
        
        instructions = [
            "",
            "Structure du fichier :",
            "- Colonne B: ID des tâches (pour référence, non importé)",
            "- Colonne C: Noms des tâches",
            "- Colonne D: Durées des tâches",
            "- Colonne E: Prédécesseurs (vide si aucun, sinon IDs séparés par virgules)",
            "- Colonne H6: Unité de temps, H7: Valeur unité (j/h/m)",
            "- Colonne H9: Temps de cycle (présent pour compatibilité, non utilisé)",
            "",
            "Format d'import attendu :",
            "- Headers en ligne 6 (C6=Tâche, D6=Durée, E6=Prédécesseur)",
            "- Données à partir de la ligne 7",
            "- Paramètres en colonne H aux lignes spécifiées",
            "",
            f"Nombre de tâches exportées: {len(tasks_data)}",
            "",
            "Note: L'algorithme Précédences ne nécessite pas de temps de cycle.",
            "Ce fichier peut être modifié et réimporté dans l'application."
        ]
        
        for i, instruction in enumerate(instructions):
            instructions_ws.cell(row=2+i, column=1, value=instruction)
        
        # Sauvegarder dans le BytesIO
        wb.save(output)
        output.seek(0)
        
        # Retourner une réponse de streaming
        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=Export_{algorithm_name}_Precedences.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'export: {str(e)}")


def export_ligne_assemblage_mixte_equilibrage_to_excel(
    products_data: List[dict],
    tasks_data: List[dict],
    cycle_time: float,
    unite: str = "minutes"
) -> bytes:
    """
    Exporte les données d'équilibrage mixte vers Excel selon le format spécifique.
    
    Structure Excel:
    - B5: "ID" (repère visuel)
    - B7+: ID des tâches (1, 2, 3...)
    - C4: "Demande" (fusionné)
    - C6: "Tâche", C7+: Noms des tâches
    - D4-E4: Demande produit 1 (fusionné)
    - D5-E5: Nom produit 1 (fusionné)
    - D6: "Durée", E6: "Prédécesseur" (pour produit 1)
    - F4-G4: Demande produit 2 (fusionné)
    - F5-G5: Nom produit 2 (fusionné)
    - F6: "Durée", G6: "Prédécesseur" (pour produit 2)
    - ... même logique pour produits 3, 4, 5
    - P5: "Unité de temps", P6: Valeur unité (j/h/m)
    - P8: "Durée de la période", P9: Valeur cycle_time
    
    Args:
        products_data: Liste des produits avec product_id, name, demand
        tasks_data: Liste des tâches avec task_id, name, times (par produit), predecessors (par produit)
        cycle_time: Durée de la période
        unite: Unité de temps
        
    Returns:
        bytes: Contenu du fichier Excel
    """
    try:
        from fastapi.responses import StreamingResponse
        
        # Validation
        if not products_data or not tasks_data:
            raise ValueError("Données de produits et tâches requises")
        
        if len(products_data) > 5:
            raise ValueError("Maximum 5 produits supportés")
        
        # Créer un BytesIO pour le fichier Excel
        output = io.BytesIO()
        
        # Créer un workbook avec openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Export_Equilibrage_Mixte"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        demand_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # B5: "ID"
        ws['B5'] = "ID"
        ws['B5'].font = header_font
        ws['B5'].fill = header_fill
        ws['B5'].alignment = Alignment(horizontal="center")
        ws['B5'].border = border
        
        # C4: "Demande" 
        ws['C4'] = "Demande"
        ws['C4'].font = demand_font
        ws['C4'].alignment = Alignment(horizontal="center")
        
        # C6: "Tâche"
        ws['C6'] = "Tâche"
        ws['C6'].font = header_font
        ws['C6'].fill = header_fill
        ws['C6'].alignment = Alignment(horizontal="center")
        ws['C6'].border = border
        
        # Configurer les colonnes pour chaque produit (max 5)
        product_columns = [
            (4, 5),   # Produit 1: colonnes D, E
            (6, 7),   # Produit 2: colonnes F, G
            (8, 9),   # Produit 3: colonnes H, I
            (10, 11), # Produit 4: colonnes J, K
            (12, 13)  # Produit 5: colonnes L, M
        ]
        
        # Remplir les données pour chaque produit
        for i, product in enumerate(products_data[:5]):  # Max 5 produits
            if i >= len(product_columns):
                break
                
            col1, col2 = product_columns[i]
            
            # Demande du produit (ligne 4, fusionnée sur 2 colonnes)
            demand_cell = ws.cell(row=4, column=col1, value=product.get("demand", 0))
            demand_cell.font = demand_font
            demand_cell.alignment = Alignment(horizontal="center")
            
            # Fusionner les cellules pour la demande
            ws.merge_cells(start_row=4, start_column=col1, end_row=4, end_column=col2)
            
            # Nom du produit (ligne 5, fusionnée sur 2 colonnes)
            name_cell = ws.cell(row=5, column=col1, value=product.get("name", f"Produit {i+1}"))
            name_cell.font = header_font
            name_cell.fill = header_fill
            name_cell.alignment = Alignment(horizontal="center")
            
            # Fusionner les cellules pour le nom
            ws.merge_cells(start_row=5, start_column=col1, end_row=5, end_column=col2)
            
            # Headers pour ce produit (ligne 6)
            # Durée
            duree_cell = ws.cell(row=6, column=col1, value="Durée")
            duree_cell.font = header_font
            duree_cell.fill = header_fill
            duree_cell.alignment = Alignment(horizontal="center")
            duree_cell.border = border
            
            # Prédécesseur
            pred_cell = ws.cell(row=6, column=col2, value="Prédécesseur")
            pred_cell.font = header_font
            pred_cell.fill = header_fill
            pred_cell.alignment = Alignment(horizontal="center")
            pred_cell.border = border
        
        # Remplir les données des tâches à partir de la ligne 7
        for i, task in enumerate(tasks_data):
            row = 7 + i
            
            # B7+: ID des tâches
            task_id = task.get("task_id", task.get("id", i + 1))
            ws.cell(row=row, column=2, value=task_id).border = border
            
            # C7+: Noms des tâches
            ws.cell(row=row, column=3, value=task.get("name", f"Tâche {task_id}")).border = border
            
            # Données pour chaque produit
            times = task.get("times", [])
            models = task.get("models", [])
            
            for j, product in enumerate(products_data[:5]):
                if j >= len(product_columns):
                    break
                    
                col1, col2 = product_columns[j]
                
                # Durée pour ce produit
                time_val = 0
                if j < len(times):
                    time_val = times[j]
                elif j < len(models) and models[j] and "time" in models[j]:
                    time_val = models[j]["time"]
                
                ws.cell(row=row, column=col1, value=time_val).border = border
                
                # Prédécesseurs pour ce produit
                pred_val = ""
                if j < len(models) and models[j] and "predecessors" in models[j]:
                    predecessors = models[j]["predecessors"]
                    if predecessors is not None:
                        if isinstance(predecessors, list):
                            pred_val = ",".join(map(str, predecessors))
                        else:
                            pred_val = str(predecessors)
                
                ws.cell(row=row, column=col2, value=pred_val).border = border
        
        # Informations supplémentaires en colonne P
        # P5: "Unité de temps"
        ws['P5'] = "Unité de temps"
        ws['P5'].font = Font(bold=True)
        
        # P6: Valeur de l'unité (j/h/m)
        unite_short = unite
        if unite.lower() == "minutes":
            unite_short = "m"
        elif unite.lower() == "heures":
            unite_short = "h"
        elif unite.lower() == "jours":
            unite_short = "j"
        ws['P6'] = unite_short
        
        # P8: "Durée de la période"
        ws['P8'] = "Durée de la période"
        ws['P8'].font = Font(bold=True)
        
        # P9: Valeur du cycle_time
        ws['P9'] = cycle_time
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 15
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws.column_dimensions[col].width = 12
        ws.column_dimensions['P'].width = 18
        
        # Ajouter un onglet d'instructions
        instructions_ws = wb.create_sheet("Instructions")
        instructions_ws['A1'] = "EXPORT ÉQUILIBRAGE MIXTE"
        instructions_ws['A1'].font = Font(bold=True, size=14)
        
        instructions = [
            "",
            "Structure du fichier :",
            "- Colonne B: ID des tâches (repère visuel)",
            "- Colonne C: Noms des tâches",
            "- Ligne 4: Demandes des produits (fusionnées sur 2 colonnes)",
            "- Ligne 5: Noms des produits (fusionnés sur 2 colonnes)",
            "- Ligne 6: Headers 'Durée' et 'Prédécesseur' pour chaque produit",
            "- À partir ligne 7: Données des tâches",
            "- Colonne P5: Unité de temps, P6: Valeur unité (j/h/m)",
            "- Colonne P8: Durée de la période, P9: Valeur",
            "",
            "Organisation par produit :",
            "- Produit 1: colonnes D-E (Durée, Prédécesseur)",
            "- Produit 2: colonnes F-G (Durée, Prédécesseur)",
            "- Produit 3: colonnes H-I (Durée, Prédécesseur)",
            "- Produit 4: colonnes J-K (Durée, Prédécesseur)",
            "- Produit 5: colonnes L-M (Durée, Prédécesseur)",
            "",
            f"Nombre de produits exportés: {len(products_data)}",
            f"Nombre de tâches exportées: {len(tasks_data)}",
            "",
            "Ce fichier peut être modifié et réimporté dans l'application."
        ]
        
        for i, instruction in enumerate(instructions):
            instructions_ws.cell(row=2+i, column=1, value=instruction)
        
        # Sauvegarder dans le BytesIO
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'export: {str(e)}")


async def parse_ligne_assemblage_mixte_equilibrage_excel(file) -> Dict:
    """
    Parse un fichier Excel pour l'équilibrage mixte selon le format spécifique.
    
    Structure attendue:
    - B5: "ID" (ignoré)
    - C4: "Demande"
    - C6: "Tâche", C7+: Noms des tâches
    - D4-E4: Demande produit 1, D5-E5: Nom produit 1, D6: "Durée", E6: "Prédécesseur"
    - F4-G4: Demande produit 2, F5-G5: Nom produit 2, F6: "Durée", G6: "Prédécesseur"
    - ... même logique pour produits 3, 4, 5
    - P5: "Unité de temps", P6: Valeur unité (j/h/m)
    - P8: "Durée de la période", P9: Valeur cycle_time
    
    Args:
        file: Fichier Excel uploadé
        
    Returns:
        Dict contenant les données formatées pour l'API
    """
    try:
        # Lire le fichier Excel
        df = pd.read_excel(file, header=None)
        
        # Vérifier la structure minimale
        if df.shape[0] < 10 or df.shape[1] < 16:
            raise ValueError("Structure de fichier incorrecte - taille insuffisante")
        
        # Extraire l'unité de temps (P6)
        unite = "minutes"  # valeur par défaut
        try:
            unite_cell = df.iloc[5, 15]  # P6 (ligne 6, colonne P)
            if pd.notna(unite_cell):
                unite_str = str(unite_cell).lower().strip()
                if unite_str == 'j':
                    unite = "jours"
                elif unite_str == 'h':
                    unite = "heures"
                elif unite_str == 'm':
                    unite = "minutes"
        except:
            pass
        
        # Extraire la durée de la période (P9)
        cycle_time = 50.0  # valeur par défaut
        try:
            cycle_time_cell = df.iloc[8, 15]  # P9 (ligne 9, colonne P)
            if pd.notna(cycle_time_cell):
                cycle_time = float(cycle_time_cell)
        except:
            pass
        
        # Extraire les données des produits
        products_data = []
        product_columns = [
            (3, 4),   # Produit 1: colonnes D, E (index 3, 4)
            (5, 6),   # Produit 2: colonnes F, G (index 5, 6)
            (7, 8),   # Produit 3: colonnes H, I (index 7, 8)
            (9, 10),  # Produit 4: colonnes J, K (index 9, 10)
            (11, 12)  # Produit 5: colonnes L, M (index 11, 12)
        ]
        
        for i, (col1, col2) in enumerate(product_columns):
            try:
                # Nom du produit (ligne 5, colonne col1)
                name_cell = df.iloc[4, col1]  # ligne 5 (index 4)
                if pd.isna(name_cell) or not str(name_cell).strip():
                    break  # Plus de produits
                
                product_name = str(name_cell).strip()
                
                # Demande du produit (ligne 4, colonne col1)
                demand_cell = df.iloc[3, col1]  # ligne 4 (index 3)
                demand = 1  # valeur par défaut
                if pd.notna(demand_cell):
                    try:
                        demand = int(float(demand_cell))
                        if demand <= 0:
                            demand = 1
                    except:
                        demand = 1
                
                products_data.append({
                    "product_id": i + 1,
                    "name": product_name,
                    "demand": demand
                })
                
            except Exception as e:
                print(f"Erreur lors de la lecture du produit {i+1}: {e}")
                break
        
        if not products_data:
            raise ValueError("Aucun produit valide trouvé")
        
        # Extraire les données des tâches
        tasks_data = []
        errors = []
        
        # Parcourir les lignes à partir de la ligne 7 (index 6)
        for i in range(6, min(df.shape[0], 50)):  # Limite à 50 tâches max
            row_num = i + 1  # Numéro de ligne Excel (1-indexé)
            
            # Nom de la tâche (colonne C)
            task_name = df.iloc[i, 2]  # colonne C (index 2)
            if pd.isna(task_name) or not str(task_name).strip():
                continue  # Ligne vide, on passe
            
            task_name = str(task_name).strip()
            
            # ID de la tâche (colonne B, ou index+1 si absent)
            task_id = df.iloc[i, 1]  # colonne B (index 1)
            if pd.isna(task_id):
                task_id = i - 5  # Calculer l'ID basé sur la position (ligne 7 = tâche 1)
            else:
                try:
                    task_id = int(float(task_id))
                except:
                    task_id = i - 5
            
            # Extraire les données pour chaque produit
            models = []
            times = []
            
            for j, (col1, col2) in enumerate(product_columns):
                if j >= len(products_data):
                    break
                
                # Durée pour ce produit (colonne col1)
                time_val = 0
                try:
                    time_cell = df.iloc[i, col1]
                    if pd.notna(time_cell):
                        time_val = float(time_cell)
                        if time_val < 0:
                            errors.append(f"Tâche '{task_name}', produit {j+1}: durée négative ({time_val})")
                            time_val = 0
                except:
                    time_val = 0
                
                times.append(time_val)
                
                # Prédécesseurs pour ce produit (colonne col2)
                predecessors = None
                try:
                    pred_cell = df.iloc[i, col2]
                    if pd.notna(pred_cell) and str(pred_cell).strip():
                        pred_str = str(pred_cell).strip()
                        if pred_str:
                            # Parser les prédécesseurs (format: "1,2,3" ou "1")
                            pred_ids = []
                            for p in pred_str.split(','):
                                try:
                                    pred_id = int(float(p.strip()))
                                    if pred_id > 0:
                                        pred_ids.append(pred_id)
                                except:
                                    pass
                            
                            if pred_ids:
                                if len(pred_ids) == 1:
                                    predecessors = pred_ids[0]
                                else:
                                    predecessors = pred_ids
                except:
                    pass
                
                models.append({
                    "predecessors": predecessors,
                    "time": time_val
                })
            
            # Ajouter la tâche
            tasks_data.append({
                "task_id": task_id,
                "id": task_id,
                "name": task_name,
                "times": times,
                "models": models
            })
        
        # Vérifier qu'on a des données valides
        if not tasks_data:
            raise ValueError("Aucune tâche valide trouvée")
        
        # Signaler les erreurs s'il y en a
        if errors:
            error_msg = "Avertissements lors de l'import:\n" + "\n".join(f"• {err}" for err in errors[:10])
            if len(errors) > 10:
                error_msg += f"\n... et {len(errors) - 10} autres avertissements"
            print(f"Import warnings: {error_msg}")
        
        return {
            "products_data": products_data,
            "tasks_data": tasks_data,
            "cycle_time": cycle_time,
            "unite": unite
        }
        
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=400, detail=f"Erreur lors de la lecture du fichier Excel: {str(e)}")
 